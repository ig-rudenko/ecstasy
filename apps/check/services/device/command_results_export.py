from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from apps.check.models import BulkDeviceCommandExecution, BulkDeviceCommandExecutionResult

EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

RESULT_HEADERS = [
    "Device ID",
    "Device name",
    "Status",
    "Command",
    "Output",
    "Detail",
    "Error",
    "Duration, sec",
    "Created at",
    "Updated at",
]


def build_bulk_command_results_workbook(
    execution: BulkDeviceCommandExecution,
    results: list[BulkDeviceCommandExecutionResult],
) -> bytes:
    """Build XLSX workbook with results for one bulk command execution."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "results"

    write_execution_summary(sheet, execution)
    write_results_table(sheet, results)
    apply_sheet_formatting(sheet)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_execution_summary(sheet: Worksheet, execution: BulkDeviceCommandExecution) -> None:
    """Write execution metadata above the results table."""
    summary_rows = [
        ("Task ID", execution.task_id),
        ("Command", execution.command_name),
        ("Status", execution.status),
        ("Progress", f"{execution.progress}%"),
        ("Processed", execution.processed),
        ("Total", execution.total),
        ("Launched at", format_excel_datetime(execution.launched_at)),
        ("Finished at", format_excel_datetime(execution.finished_at)),
    ]

    for row_index, (label, value) in enumerate(summary_rows, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=value)


def write_results_table(sheet: Worksheet, results: list[BulkDeviceCommandExecutionResult]) -> None:
    """Write result rows to the worksheet."""
    header_row = 10
    for column_index, header in enumerate(RESULT_HEADERS, start=1):
        sheet.cell(row=header_row, column=column_index, value=header)

    for row_index, result in enumerate(results, start=header_row + 1):
        sheet.cell(row=row_index, column=1, value=result.device_id)
        sheet.cell(row=row_index, column=2, value=result.device_name)
        sheet.cell(row=row_index, column=3, value=result.status)
        sheet.cell(row=row_index, column=4, value=result.command_text)
        sheet.cell(row=row_index, column=5, value=result.output)
        sheet.cell(row=row_index, column=6, value=result.detail)
        sheet.cell(row=row_index, column=7, value=result.error)
        sheet.cell(row=row_index, column=8, value=result.duration)
        sheet.cell(row=row_index, column=9, value=format_excel_datetime(result.created_at))
        sheet.cell(row=row_index, column=10, value=format_excel_datetime(result.updated_at))


def format_excel_datetime(value) -> str:
    """Return a timezone-aware datetime as a stable Excel cell string."""
    if not value:
        return ""
    return timezone.localtime(value).strftime("%Y-%m-%d %H:%M:%S")


def apply_sheet_formatting(sheet: Worksheet) -> None:
    """Apply minimal formatting for readable command output cells."""
    sheet.freeze_panes = "A11"
    for cell in sheet[10]:
        cell.font = Font(bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 12,
        "B": 28,
        "C": 14,
        "D": 36,
        "E": 60,
        "F": 40,
        "G": 40,
        "H": 14,
        "I": 22,
        "J": 22,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
